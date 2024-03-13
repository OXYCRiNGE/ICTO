import pandas as pd
import string
import re
import openpyxl
from openpyxl.styles import PatternFill
from decimal import *
from datetime import datetime
import os
from math import trunc

path = os.path.abspath(os.curdir)

def check_file(path, file_name):
    check = os.path.exists(f'{path}\{file_name}')
    if not check:
        print(f'Файл {file_name} отсутствует в папке по пути {path}')
        input('Нажмите enter чтобы выйти...')
        exit()

check_file(path, "Справочники.xlsx")
check_file(path, "2 Шаблон Родитель.xlsx")

sheet_name_spravochnik = [0, 1, 3, 4]
parents = []
gender = []
passport_type = []
citizenship = []

for i in range(len(sheet_name_spravochnik)):
    list = pd.read_excel('Справочники.xlsx', sheet_name=sheet_name_spravochnik[i])
    match i:
        case 0:
            for i in range(len(list)):
                parents.append(list['name'][i])
        case 1:
            for i in range(len(list)):
                gender.append(list['title'][i])
        case 2:
            for i in range(len(list)):
                passport_type.append(list['name'][i])
        case 3:
            for i in range(len(list)):
                citizenship.append(list['title'][i])

df = pd.read_excel('2 Шаблон Родитель.xlsx')
wb = openpyxl.load_workbook('2 Шаблон Родитель.xlsx')
df_column_list = df.columns


def change_color(wb, row, col):
    ws = wb['Контингент Родители']
    cell = ws.cell(row+2, col+1)
    fill = PatternFill(start_color='FF0000',
                       end_color='FF0000', fill_type='solid')
    cell.fill = fill


def is_allowed(data_string, cell):
    allowed = set(data_string)
    if set(str(cell)) & allowed:
        pass
    else:
        change_color(wb, i, j)


def is_disallowed(data_string, cell):
    disallowed = set(data_string)
    if set(str(cell)) & disallowed:
        change_color(wb, i, j)


def find_regex(value, cell):
    regex = re.compile(value)
    search = regex.search(str(cell))
    if search == None:
        change_color(wb, i, j)


def find_in_list(list, cell):
    if cell not in list:
        change_color(wb, i, j)


def snils_number(snilsNumber):
    # print('Цифры номера:', snilsNumber)
    i = 9
    prin_i = ''
    s = 0
    prin_s = ''
    for x in range(i):
        prin_i += str(i) + ' '
        s += int(snilsNumber[x]) * i
        prin_s += snilsNumber[x] + '*' + str(i) + '+'
        i -= 1
    prin_s = prin_s[:-1]
    return s

def is_valid(pattern, cell):
    return bool(re.match(pattern, cell))


def contains_only_numbers(string):
    # Проверяем, что строка содержит только цифры и/или точки
    if not all(char.isdigit() or char == '.' for char in string):
        return False

    # Проверяем, что строка может быть преобразована в число
    try:
        float(string)
        return True
    except ValueError:
        return False


def check_date_format_passport(date_str):
    pattern1 = r'^\d{2}-\d{2}-\d{4}$'
    pattern2 = r'^\d{2}\.\d{2}\.\d{4}$'
    pattern3 = r'^\d{4}-\d{2}-\d{2}\s\d{2}:\d{2}:\d{2}$'
    if re.match(pattern1, date_str):
        date_format = '%d-%m-%Y'
    elif re.match(pattern2, date_str):
        date_format = '%d.%m.%Y'
    elif re.match(pattern3, date_str):
        date_format = '%Y-%m-%d %H:%M:%S'
    else:
        change_color(wb, i, j)
        return False
    try:
        datetime.strptime(date_str, date_format)
        return True
    except ValueError:
        change_color(wb, i, j)
        return False
    

def check_date_format_dof(date_str):
    pattern1 = r'^\d{2}-\d{2}-\d{4}$'
    pattern2 = r'^\d{2}\.\d{2}\.\d{4}$'
    pattern3 = r'^\d{4}-\d{2}-\d{2}\s\d{2}:\d{2}:\d{2}$'
    if re.match(pattern1, date_str):
        date_format = '%d-%m-%Y'
    elif re.match(pattern2, date_str):
        date_format = '%d.%m.%Y'
    elif re.match(pattern3, date_str):
        date_format = '%Y-%m-%d %H:%M:%S'
    else:
        change_color(wb, i, j)
        return False
    try:
        dob = datetime.strptime(date_str, date_format)
        now = datetime.now()
        age = now.year - dob.year - ((now.month, now.day) < (dob.month, dob.day))
        if age < 18:
            change_color(wb, i, j)
            return False
        elif dob.year % 4 == 0 and (dob.year % 100 != 0 or dob.year % 400 == 0):
            return True, age, True
        else:
            return True, age, False
    except ValueError:
        change_color(wb, i, j)
        return False


for i in range(len(df)):
    # print('----------------')
    for j in range(len(df_column_list)):
        cell = df[df_column_list[j]][i]
        match j:

            ## Внутренний идентификатор РЕБЕНКА и РОДИТЕЛЯ ----------------------------
            case (0 | 1):
                if pd.isnull(cell):
                    change_color(wb, i, j)
                    continue
                if type(cell) != str:
                    cell = trunc(cell)
                if not contains_only_numbers(str(cell)):
                    change_color(wb, i, j)
               

            # Тип родителя -----------------------------------------------------------
            case 2:
                if pd.isnull(cell):
                    change_color(wb, i, j)
                    continue
                find_in_list(parents, cell)


            # # Фамилия, имя --------------------------------------------------------
            case (3 | 4):
                if pd.isnull(cell):
                    change_color(wb, i, j)
                    continue
                is_disallowed(string.printable, cell)
            
            # # Отчество -------------------------------------------------------------------
            case (5):
                if pd.isnull(cell):
                    continue
                is_disallowed(string.printable, cell)

            # # Дата рождения ------------------------------------------------------------
            case (6):
                if pd.isnull(cell):
                    change_color(wb, i, j)
                    continue
                date_str = str(cell)
                check_date_format_dof(date_str)

            # Пол --------------------------------------------------------------------
            case 7:
                if pd.isnull(cell):
                    change_color(wb, i, j)
                    continue
                find_in_list(gender, cell)

            # Снилс -----------------------------------------------------------------
            case 8:
                if pd.isnull(cell):
                    change_color(wb, i, j)
                    continue
                regex = '\d{3}-\d{3}-\d{3}\s\d{2}'
                find_regex(regex, cell)

                if len(str(cell)) != 14:
                    continue
                
                cell = cell.replace('-', '')
                snilsNumber = cell[:9]
                snilsEnd = int(cell[10:12])

                number_end = snils_number(snilsNumber)

                while number_end >= 101:
                    number_end = number_end - 101

                if number_end < 100 and number_end == snilsEnd:
                    continue
                elif number_end == 100 or number_end == 101:
                    continue
                else:
                    change_color(wb, i, j)
                    


            # Паспорт серия ------------------------------------------------------------------
            case (9):
                if pd.isnull(cell):
                    change_color(wb, i, j)
                    continue
                cell = trunc(cell)
                if not contains_only_numbers(str(cell)) or len(str(cell)) != 4:
                    change_color(wb, i, j)

            # Паспорт номер -------------------------------------------------------------------
            case (10):
                # print(cell)
                if pd.isnull(cell):
                    change_color(wb, i, j)
                    continue
                if type(cell) != str:
                    cell = trunc(cell)
                if not contains_only_numbers(str(cell)) or len(str(cell)) != 6:
                    change_color(wb, i, j)

            # Паспорт тип ------------------------------------------------------------------
            case 11:
                if pd.isnull(cell):
                    change_color(wb, i, j)
                    continue
                find_in_list(passport_type, cell)

            # # Паспорт дата выдачи -----------------------------------------------------------
            case 12:
                if pd.isnull(cell):
                    continue
                date_str = str(cell)
                check_date_format_passport(date_str)

            # Паспорт кем выдано -------------------------------------------------------------
            case 13:
                if pd.isnull(cell):
                    continue
                data_string = string.ascii_letters + string.digits + string.punctuation
                is_disallowed(data_string, cell)

            # Гражданство --------------------------------------------------------------------
            case 14:
                if pd.isnull(cell):
                    change_color(wb, i, j)
                    continue
                find_in_list(citizenship, cell)

            # Телефон ------------------------------------------------------------------------
            case 15:
                if pd.isnull(cell):
                    change_color(wb, i, j)
                    continue
                data_string = string.ascii_letters + string.punctuation + string.whitespace
                data_string = data_string.replace(';', '')
                is_disallowed(data_string, cell)
                cell = str(cell).split(';')
                for k in range(len(cell)):
                    cell[k] = cell[k].replace(' ', '')
                    if len(cell[k]) != 10:
                        change_color(wb, i, j)
                        break
            
            # Email --------------------------------------------------------------------------
            case 16:
                if pd.isnull(cell):
                    change_color(wb, i, j)
                    continue
                pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
                if not is_valid(pattern, cell):
                    change_color(wb, i, j)


wb.save('2 Шаблон Родитель.xlsx')